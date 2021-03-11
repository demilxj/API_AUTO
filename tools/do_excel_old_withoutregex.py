# -*- coding: utf-8 -*-
# @Time : 2021/3/2 16:18
# @Author : demi
# @Email : demilxj@foxmail.com
# @File : do_excel.py

from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools import project_path


class DoExcel:
    loan_id = None
    @classmethod
    def get_data(cls, file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(project_path.test_config_path, 'MODE', 'mode'))


        tel = cls.get_tel(project_path.test_case_path, 'init', 2, 1)  # 获取excel中的值
        normal_tel = cls.get_tel(project_path.test_case_path, 'init', 3, 1)
        admin_tel = cls.get_tel(project_path.test_case_path, 'init', 4, 1)
        loan_member_id = cls.get_tel(project_path.test_case_path, 'init', 5, 1)
        memberID = cls.get_tel(project_path.test_case_path, 'init', 6, 1)

        # 利用python查询数据库的方式，去拿到最大的手机号——可以加到这里，也可以加到get_data里面
        test_data = []
        for key in mode:  # 遍历这个存在配置文件里面的字典
            sheet = wb[key]  # key是表单名sheet_name
            if mode[key] == 'all':
                for i in range(2, sheet.max_row+1):
                    row_data = {}
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['url'] = sheet.cell(i, 2).value
                    # row_data['data'] = sheet.cell(i, 3).value
                    # 做手机号的数据替换
                    if sheet.cell(i, 3).value.find('${tel}') != -1:  # 有找到这个字符串${tel}
                        row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel))
                        tel = tel + 1  # 每次完成tel调用后 就加1
                    elif sheet.cell(i, 3).value.find('${normal_tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${normal_tel}', str(normal_tel))
                    elif sheet.cell(i, 3).value.find('${admin_tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${admin_tel}', str(admin_tel))
                    elif sheet.cell(i, 3).value.find('${loan_member_id}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${loan_member_id}', str(loan_member_id))
                    elif sheet.cell(i, 3).value.find('${memberID}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${memberID}', str(memberID))
                    else:  # 没有找到的话
                        row_data['data'] = sheet.cell(i, 3).value
                    if sheet.cell(i, 4).value != None:
                        if sheet.cell(i, 4).value.find('${normal_tel}') != -1:
                            row_data['check_sql'] = sheet.cell(i, 4).value.replace('${normal_tel}',
                                                                                             str(normal_tel))
                    else:
                        row_data['check_sql'] = sheet.cell(i, 4).value
                    row_data['title'] = sheet.cell(i, 5).value
                    row_data['http_method'] = sheet.cell(i, 6).value
                    row_data['expected'] = sheet.cell(i, 7).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    cls.update_tel(project_path.test_case_path, 'init', tel+2)  # 更新excel中的手机号数据
            else:
                for case_id in mode[key]:
                    row_data = {}
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    # row_data['data'] = sheet.cell(case_id+1, 3).value
                    if sheet.cell(case_id+1, 3).value.find('${tel}') != -1:  # 有找到这个字符串${tel}
                        row_data['data'] = sheet.cell(case_id+1, 3).value.replace('${tel}', str(tel))
                        tel = tel + 1
                    elif sheet.cell(case_id+1, 3).value.find('${normal_tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(normal_tel))
                    elif sheet.cell(case_id+1, 3).value.find('${admin_tel}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${admin_tel}', str(admin_tel))
                    elif sheet.cell(case_id+1, 3).value.find('${loan_member_id}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${loan_member_id}', str(loan_member_id))
                    elif sheet.cell(case_id+1, 3).value.find('${memberID}') != -1:
                        row_data['data'] = sheet.cell(i, 3).value.replace('${memberID}', str(memberID))
                    else:  # 没有找到的话
                        row_data['data'] = sheet.cell(case_id+1, 3).value
                    if sheet.cell(case_id+1, 4).value != None:
                        if sheet.cell(case_id+1, 4).value.find('${normal_tel}') != -1:
                            row_data['check_sql'] = sheet.cell(case_id + 1, 4).value.replace('${normal_tel}', str(normal_tel))
                    else:
                         row_data['check_sql'] = sheet.cell(case_id + 1, 4).value
                    row_data['title'] = sheet.cell(case_id+1, 5).value
                    row_data['http_method'] = sheet.cell(case_id+1, 6).value
                    row_data['expected'] = sheet.cell(case_id+1, 7).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
                    cls.update_tel(project_path.test_case_path, 'init', tel)  # 更新excel中的手机号数据
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, row, col, result):  # 专门写回数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(row, col).value = result
        wb.save(file_name)  # 保存结果

    @classmethod
    def get_tel(cls, file_name, sheet_name, i, j):  # 获取excel中的手机号
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        tel = sheet.cell(i, j).value
        return tel

    @classmethod
    def update_tel(cls, file_name, sheet_name, tel):  # 更新excel中的手机号数据
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(2, 1).value = tel
        wb.save(file_name)


if __name__ == '__main__':
    test_data = DoExcel().get_data(project_path.test_case_path)
    print(test_data)


