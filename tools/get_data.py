# -*- coding: utf-8 -*-
# @Time : 2021/3/3 10:43
# @Author : demi
# @Email : demilxj@foxmail.com
# @File : get_data.py

from tools import project_path
# import pandas as pd
from openpyxl import load_workbook
from tools.read_config import ReadConfig


class Gettel:
    @staticmethod
    def get_tel(file_name, sheet_name, i, j):  # 获取excel中的手机号
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        tel = sheet.cell(i, j).value
        return tel

class GetData:
    Cookie = None
    loan_id = None
    check_list = eval(ReadConfig().get_config(project_path.test_config_path, 'CHECKLEAVEAMOUNT', 'check_list'))
    # NoRegTel = pd.read_excel(test_case_path, sheet_name='init').ix[0,0]  # 利用pandas取值
    # normal_tel = pd.read_excel(test_case_path, sheet_name='init').ix[1,0]  # 利用pandas取值
    # admin_tel = pd.read_excel(test_case_path, sheet_name='init').ix[2,0]  # 利用pandas取值
    # loan_member_id = pd.read_excel(test_case_path, sheet_name='init').ix[3,0]  # 利用pandas取值
    # memberID = pd.read_excel(test_case_path, sheet_name='init').ix[4,0]  # 利用pandas取值

    NoRegTel = Gettel.get_tel(project_path.test_case_path, 'init', 2, 1)  # 获取excel中的值
    normal_tel = Gettel.get_tel(project_path.test_case_path, 'init', 3, 1)
    admin_tel = Gettel.get_tel(project_path.test_case_path, 'init', 4, 1)
    loan_member_id = Gettel.get_tel(project_path.test_case_path, 'init', 5, 1)
    memberID = Gettel.get_tel(project_path.test_case_path, 'init', 6, 1)

if __name__ == '__main__':
        # print(GetData().loan_member_id)
        # print(type(GetData().loan_member_id))
        print(GetData().check_list)
        print(GetData().memberID)
        print(type(GetData().memberID))
